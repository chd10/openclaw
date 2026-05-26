import os
import json
import fcntl
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

LOG_FILE  = "/data/emails_log.json"
XLSX_FILE = "/data/emails_log.xlsx"
_LOCK_FILE = "/data/emails_log.lock"

_HEADERS = [
    ("id",             "Номер",           8),
    ("email",          "Email",           32),
    ("subject",        "Тема",            40),
    ("date",           "Дата отправки",   18),
    ("delivered",      "Доставлен",       12),
    ("opened",         "Открыт",          10),
    ("opened_time",    "Время открытия",  18),
    ("reply",          "Ответ",           10),
    ("reply_date",     "Дата ответа",     18),
    ("campaign",       "Кампания",        16),
    ("template_variant","Шаблон",          9),
]

_FILL_ODD  = PatternFill("solid", fgColor="FFFFFF")
_FILL_EVEN = PatternFill("solid", fgColor="F0F4FA")
_FONT_HEAD = Font(bold=True, color="FFFFFF")
_FILL_HEAD = PatternFill("solid", fgColor="2E6DB4")


def _fmt_bool(val):
    if val is True or val == "да":
        return "Да"
    if val is False or val == "нет":
        return "Нет"
    return str(val) if val is not None else ""


def _fmt_date(val):
    if not val:
        return ""
    try:
        dt = datetime.fromisoformat(str(val))
        return dt.strftime("%d.%m.%Y %H:%M")
    except Exception:
        return str(val)[:16]


def _fmt_delivered(val):
    mapping = {"pending": "Pending", "bounce": "Bounce", "да": "Да"}
    return mapping.get(str(val), str(val)) if val else ""


def _load():
    if os.path.exists(LOG_FILE):
        with open(LOG_FILE) as f:
            return json.load(f)
    return []


def _save(records):
    with open(LOG_FILE, "w") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)
    _write_xlsx(records)


def _write_xlsx(records):
    wb = Workbook()
    ws = wb.active
    ws.title = "Email Log"

    # Headers
    for col, (_, title, width) in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = _FONT_HEAD
        cell.fill = _FILL_HEAD
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 20

    # Rows
    _DATE_COLS  = {"date", "opened_time", "reply_date"}
    _BOOL_COLS  = {"opened", "reply"}
    _DELIV_COLS = {"delivered"}

    for row_idx, r in enumerate(records, start=2):
        fill = _FILL_EVEN if row_idx % 2 == 0 else _FILL_ODD
        for col, (field, _, _) in enumerate(_HEADERS, start=1):
            raw = r.get(field)
            if field in _DATE_COLS:
                val = _fmt_date(raw)
            elif field in _BOOL_COLS:
                val = _fmt_bool(raw)
            elif field in _DELIV_COLS:
                val = _fmt_delivered(raw)
            else:
                val = raw if raw is not None else ""
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = fill
            cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"
    wb.save(XLSX_FILE)


def _with_lock(fn):
    os.makedirs("/data", exist_ok=True)
    with open(_LOCK_FILE, "w") as lf:
        fcntl.flock(lf, fcntl.LOCK_EX)
        try:
            return fn()
        finally:
            fcntl.flock(lf, fcntl.LOCK_UN)


def append_entry(email, subject, campaign, template_variant, track_token=None):
    def _do():
        records = _load()
        entry = {
            "id": len(records) + 1,
            "email": email,
            "subject": subject,
            "date": datetime.now().isoformat(timespec="seconds"),
            "delivered": "pending",
            "opened": False,
            "opened_time": None,
            "reply": False,
            "reply_date": None,
            "campaign": campaign,
            "template_variant": template_variant,
            "track_token": track_token,
        }
        records.append(entry)
        _save(records)
        return entry["id"]
    return _with_lock(_do)


def update_opened(track_token, opened_time=None):
    def _do():
        records = _load()
        for r in records:
            if r.get("track_token") == track_token and r["opened"] in (False, "нет"):
                r["opened"] = True
                r["opened_time"] = opened_time or datetime.now().isoformat(timespec="seconds")
                _save(records)
                return True
        return False
    return _with_lock(_do)


def update_delivered_bounce(email):
    def _do():
        records = _load()
        for r in reversed(records):
            if r.get("email", "").lower() == email.lower() and r.get("delivered") in ("pending", "да"):
                r["delivered"] = "bounce"
                _save(records)
                return True
        return False
    return _with_lock(_do)


def update_reply(email, reply_date=None):
    def _do():
        records = _load()
        for r in reversed(records):
            if r.get("email", "").lower() == email.lower() and r.get("reply") in (False, "нет"):
                r["reply"] = True
                r["reply_date"] = reply_date or datetime.now().isoformat(timespec="seconds")
                _save(records)
                return True
        return False
    return _with_lock(_do)


def get_last_n(n=10):
    return _load()[-n:]


def get_xlsx_path():
    return XLSX_FILE
